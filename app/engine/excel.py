"""Excel diff engine — cell-level comparison of spreadsheets using pandas."""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from enum import StrEnum
from typing import Any

import pandas as pd


class CellChangeType(StrEnum):
    EQUAL = "equal"
    MODIFIED = "modified"
    ADDED = "added"
    REMOVED = "removed"


@dataclass(frozen=True)
class CellDiff:
    """Diff result for a single cell."""

    row: int  # 0-indexed
    col: str  # column name
    change_type: CellChangeType
    value_a: Any = None
    value_b: Any = None


@dataclass(frozen=True)
class SheetDiff:
    """Diff result for a single sheet."""

    sheet_name: str
    columns_a: list[str]
    columns_b: list[str]
    row_count_a: int
    row_count_b: int
    cells: list[CellDiff] = field(default_factory=list)

    @property
    def added_cols(self) -> list[str]:
        return [c for c in self.columns_b if c not in self.columns_a]

    @property
    def removed_cols(self) -> list[str]:
        return [c for c in self.columns_a if c not in self.columns_b]

    @property
    def added_rows(self) -> int:
        return max(0, self.row_count_b - self.row_count_a)

    @property
    def removed_rows(self) -> int:
        return max(0, self.row_count_a - self.row_count_b)

    @property
    def modified_cells(self) -> int:
        return sum(1 for c in self.cells if c.change_type == CellChangeType.MODIFIED)

    @property
    def has_diff(self) -> bool:
        return bool(self.cells) or self.columns_a != self.columns_b


@dataclass(frozen=True)
class ExcelDiffResult:
    """Result of comparing two Excel/CSV files."""

    sheets: list[SheetDiff] = field(default_factory=list)

    @property
    def changed_sheets(self) -> int:
        return sum(1 for s in self.sheets if s.has_diff)

    @property
    def unchanged_sheets(self) -> int:
        return len(self.sheets) - self.changed_sheets


def _read_excel_sheets(file_bytes: bytes, filename: str = "") -> dict[str, pd.DataFrame]:
    """Read all sheets from an Excel or CSV file into DataFrames.

    Args:
        file_bytes: Raw file bytes.
        filename: Original filename (used to detect CSV vs Excel).

    Returns:
        Dict mapping sheet name to DataFrame.
    """
    is_csv = filename.lower().endswith(".csv") if filename else False

    if is_csv or (len(file_bytes) > 0 and not file_bytes.startswith(b"PK")):
        # Try CSV parsing
        try:
            df = pd.read_csv(io.BytesIO(file_bytes), dtype=str)
            return {"Sheet1": df}
        except Exception:  # noqa: BLE001
            pass

    # Excel (xlsx) — read via openpyxl for consistent multi-sheet support
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    result: dict[str, pd.DataFrame] = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.values)  # type: ignore[union-attr]
        if not rows:
            result[sheet_name] = pd.DataFrame()
            continue
        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
        data_rows = [[str(v) if v is not None else "" for v in row] for row in rows[1:]]
        if data_rows:
            result[sheet_name] = pd.DataFrame(data_rows, columns=headers)
        else:
            result[sheet_name] = pd.DataFrame(columns=headers)
    wb.close()
    return result


def read_excel_sheets(file_bytes: bytes, filename: str = "") -> dict[str, pd.DataFrame]:
    """Public API: read all sheets from an Excel/CSV file.

    Args:
        file_bytes: Raw file bytes.
        filename: Original filename for format detection.

    Returns:
        Dict mapping sheet name to pandas DataFrame.
    """
    return _read_excel_sheets(file_bytes, filename)


def diff_dataframes(
    df_a: pd.DataFrame,
    df_b: pd.DataFrame,
    sheet_name: str = "Sheet1",
) -> SheetDiff:
    """Compare two DataFrames cell by cell.

    Compares rows up to min(len(df_a), len(df_b)) and notes extra rows as
    ADDED or REMOVED. Compares only columns present in both; marks extra
    columns as ADDED/REMOVED via SheetDiff metadata.

    Args:
        df_a: Original DataFrame.
        df_b: Modified DataFrame.
        sheet_name: Name of the sheet being compared.

    Returns:
        SheetDiff with per-cell change information.
    """
    cols_a = list(df_a.columns)
    cols_b = list(df_b.columns)
    common_cols = [c for c in cols_a if c in cols_b]

    rows_a = len(df_a)
    rows_b = len(df_b)
    compared_rows = min(rows_a, rows_b)

    cells: list[CellDiff] = []

    # Compare common columns row by row
    for col in common_cols:
        series_a = df_a[col].astype(str)
        series_b = df_b[col].astype(str)

        for row_idx in range(compared_rows):
            val_a = series_a.iloc[row_idx]
            val_b = series_b.iloc[row_idx]
            if val_a != val_b:
                cells.append(
                    CellDiff(
                        row=row_idx,
                        col=col,
                        change_type=CellChangeType.MODIFIED,
                        value_a=val_a,
                        value_b=val_b,
                    )
                )

        # Extra rows in A (removed)
        for row_idx in range(compared_rows, rows_a):
            cells.append(
                CellDiff(
                    row=row_idx,
                    col=col,
                    change_type=CellChangeType.REMOVED,
                    value_a=series_a.iloc[row_idx],
                )
            )

        # Extra rows in B (added)
        for row_idx in range(compared_rows, rows_b):
            cells.append(
                CellDiff(
                    row=row_idx,
                    col=col,
                    change_type=CellChangeType.ADDED,
                    value_b=series_b.iloc[row_idx],
                )
            )

    return SheetDiff(
        sheet_name=sheet_name,
        columns_a=cols_a,
        columns_b=cols_b,
        row_count_a=rows_a,
        row_count_b=rows_b,
        cells=cells,
    )


def diff_excel(
    file_a: bytes,
    file_b: bytes,
    filename_a: str = "",
    filename_b: str = "",
) -> ExcelDiffResult:
    """Compare two Excel/CSV files sheet by sheet.

    Compares sheets that exist in both files. Extra sheets in either file
    are noted via SheetDiff with all rows marked ADDED or REMOVED.

    Args:
        file_a: Raw bytes of the original file.
        file_b: Raw bytes of the modified file.
        filename_a: Original filename for format detection.
        filename_b: Modified filename for format detection.

    Returns:
        ExcelDiffResult with per-sheet comparison data.
    """
    sheets_a = _read_excel_sheets(file_a, filename_a)
    sheets_b = _read_excel_sheets(file_b, filename_b)

    all_sheets = list(dict.fromkeys(list(sheets_a.keys()) + list(sheets_b.keys())))
    sheet_diffs: list[SheetDiff] = []

    for sheet_name in all_sheets:
        if sheet_name in sheets_a and sheet_name in sheets_b:
            sheet_diffs.append(diff_dataframes(sheets_a[sheet_name], sheets_b[sheet_name], sheet_name))
        elif sheet_name in sheets_a:
            # Sheet removed entirely
            df = sheets_a[sheet_name]
            cells = [
                CellDiff(row=r, col=c, change_type=CellChangeType.REMOVED, value_a=str(df[c].iloc[r]))
                for r in range(len(df))
                for c in df.columns
            ]
            sheet_diffs.append(
                SheetDiff(
                    sheet_name=sheet_name,
                    columns_a=list(df.columns),
                    columns_b=[],
                    row_count_a=len(df),
                    row_count_b=0,
                    cells=cells,
                )
            )
        else:
            # Sheet added
            df = sheets_b[sheet_name]
            cells = [
                CellDiff(row=r, col=c, change_type=CellChangeType.ADDED, value_b=str(df[c].iloc[r]))
                for r in range(len(df))
                for c in df.columns
            ]
            sheet_diffs.append(
                SheetDiff(
                    sheet_name=sheet_name,
                    columns_a=[],
                    columns_b=list(df.columns),
                    row_count_a=0,
                    row_count_b=len(df),
                    cells=cells,
                )
            )

    return ExcelDiffResult(sheets=sheet_diffs)
